#!/usr/bin/env python3
"""
plot_rarity_diff_grid.py

Same KO x genome grid layout as plot_pathway_ko_grid.py (pathway color
bars above, KO labels below, row labels to the right), but each cell
is colored by how the genome's presence/absence of that KO compares
to its genus background rate (from mean_behavior.xlsx, which records
what fraction of GlobDB genomes in the same genus carry each KO).

Per-cell classification (gene in {0,1} from summary.xlsx 'Genes' sheet,
genus_mean in [0,1] from mean_behavior.xlsx):

    gene=1, genus_mean <= --rare-threshold   -> RARE GAIN   (green)
        genome has a KO that's rare in its genus
    gene=0, genus_mean >= --common-threshold -> RARE LOSS   (red)
        genome lacks a KO that's common in its genus
    gene=1, otherwise                        -> normal presence (grey)
    gene=0, otherwise                        -> normal absence (blank)

Rows where genus == "unknown" have no GlobDB reference at all, so
genus_mean is meaninglessly 0 everywhere (not genuinely "rare"). Those
rows are grey-hatched instead of colored.

Usage
-----
python scripts/plot_rarity_diff_grid.py \\
    --summary        runs/06032026_5PM/summary.xlsx \\
    --mean-behavior   runs/06032026_5PM/mean_behavior.xlsx \\
    --output          runs/06032026_5PM/rarity_grid.pdf \\
    --title           "Genome vs. genus background"
"""

import argparse
import re
import sys
import textwrap
from pathlib import Path

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from plot_pathway_ko_grid import (
    DEFAULT_IGNORE,
    _PALETTE,
    _PATHWAY_COLORS,
    load_ko_names,
    resolve_display_label,
)

META_COLS = 5  # genome, species, genus, family, full_lineage

# Cell categories
NORMAL_ABSENT  = 0
NORMAL_PRESENT = 1
RARE_GAIN      = 2
RARE_LOSS      = 3

CATEGORY_COLORS = {
    NORMAL_ABSENT:  '#ffffff',  # nothing
    NORMAL_PRESENT: '#bbbbbb',  # grey
    RARE_GAIN:      '#2ca02c',  # green
    RARE_LOSS:      '#d62728',  # red
}


def _strip_dedup_suffix(name):
    """Undo pandas' automatic '.1', '.2', ... suffixing of duplicate column headers."""
    return re.sub(r'\.\d+$', '', name)


def load_numeric_grid(path, sheet_name):
    """
    Reads a sheet shaped like summary.xlsx's 'Genes' sheet (two header
    rows: pathway names then KO IDs, META_COLS phylogeny columns, then
    one numeric column per (pathway, KO)).

    Returns
    -------
    genomes        : list of genome IDs in row order
    species_labels : {genome_id: species_label}
    genus_labels   : {genome_id: genus_label}
    lineages       : {genome_id: full_lineage_string}
    ko_columns     : list of (pathway, ko) in sheet order
    values         : {(genome_id, pathway, ko): float}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No '{sheet_name}' sheet found in {path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        raise ValueError(f"'{sheet_name}' sheet in {path} is missing headers or data.")

    header0 = rows[0]
    header1 = rows[1]

    ko_columns = []
    last_pathway = None
    for i in range(META_COLS, len(header0)):
        h0 = header0[i]
        h1 = header1[i] if i < len(header1) else None
        if h0 is not None and str(h0).strip():
            last_pathway = _strip_dedup_suffix(str(h0).strip())
        ko = str(h1).strip() if h1 is not None and str(h1).strip() else ''
        ko_columns.append((last_pathway or '', ko))

    genomes = []
    species_labels = {}
    genus_labels = {}
    lineages = {}
    values = {}

    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        genome_id = str(row[0]).strip()
        species = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        genus = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        full_lineage = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
        genomes.append(genome_id)
        species_labels[genome_id] = species
        genus_labels[genome_id] = genus
        lineages[genome_id] = full_lineage

        for i, (pathway, ko) in enumerate(ko_columns):
            if not pathway or not ko:
                continue
            col_idx = META_COLS + i
            cell = row[col_idx] if col_idx < len(row) else None
            try:
                v = float(cell) if cell is not None else 0.0
            except (TypeError, ValueError):
                v = 0.0
            values[(genome_id, pathway, ko)] = v

    return genomes, species_labels, genus_labels, lineages, ko_columns, values


def find_empty_pathways(genomes, values_genes, ko_columns, pathway_order):
    """Pathways with no KO present in any genome (nothing to show)."""
    empty = []
    for pw in pathway_order:
        kos = [ko for p, ko in ko_columns if p == pw]
        any_present = any(
            values_genes.get((gid, pw, ko), 0.0) > 0
            for gid in genomes for ko in kos
        )
        if not any_present:
            empty.append(pw)
    return empty


def build_category_grid(
    genomes, species_labels, genus_labels, lineages,
    ko_columns_genes, values_genes, ko_columns_mean, values_mean,
    pathway_order, rare_threshold, common_threshold,
):
    """
    Returns
    -------
    ko_columns : list of (pathway, ko) used, filtered to pathway_order
                 and present in both files
    matrix     : ndarray (n_genomes x n_kos) of category codes
    row_labels : display label per genome row
    no_ref_row : bool ndarray (n_genomes,), True if genus == 'unknown'
    """
    mean_set = set(ko_columns_mean)
    ko_columns = [c for c in ko_columns_genes if c[0] in pathway_order and c in mean_set]

    n_rows = len(genomes)
    n_cols = len(ko_columns)
    matrix = np.zeros((n_rows, n_cols), dtype=int)
    row_labels = []
    no_ref_row = np.zeros(n_rows, dtype=bool)

    for r, gid in enumerate(genomes):
        sp = species_labels.get(gid, gid)
        row_labels.append(resolve_display_label(sp, lineages.get(gid, ''), gid))
        no_ref_row[r] = genus_labels.get(gid, '').lower() in ('', 'unknown')

        for c, (pathway, ko) in enumerate(ko_columns):
            gene = values_genes.get((gid, pathway, ko), 0.0)
            mean = values_mean.get((gid, pathway, ko), 0.0)
            if gene > 0 and mean <= rare_threshold:
                matrix[r, c] = RARE_GAIN
            elif gene <= 0 and mean >= common_threshold:
                matrix[r, c] = RARE_LOSS
            elif gene > 0:
                matrix[r, c] = NORMAL_PRESENT
            else:
                matrix[r, c] = NORMAL_ABSENT

    return ko_columns, matrix, row_labels, no_ref_row


def plot_category_grid(ko_columns, matrix, row_labels, no_ref_row, title, output_path,
                        ko_names=None, rare_threshold=0.10, common_threshold=0.90):
    n_rows, n_cols = matrix.shape

    pathway_order, seen = [], set()
    for pw, _ in ko_columns:
        if pw not in seen:
            pathway_order.append(pw)
            seen.add(pw)
    _fallback = iter(c for c in _PALETTE if c not in _PATHWAY_COLORS.values())
    pw_color = {p: _PATHWAY_COLORS.get(p) or next(_fallback) for p in pathway_order}

    # ── geometry (matches plot_pathway_ko_grid) ────────────────────────────
    cell_w     = 1.1
    cell_h     = 0.55
    label_w    = max((len(s) for s in row_labels), default=10) * 0.18 + 0.5
    ko_label_h = 5.0
    pw_bar_h   = 0.45
    pw_label_h = 3.0
    legend_h   = 0.9

    grid_w = n_cols * cell_w
    grid_h = n_rows * cell_h

    fig_w = max(14, label_w + grid_w + 0.5)
    fig_h = ko_label_h + grid_h + pw_bar_h + pw_label_h + legend_h + 0.3

    fig = plt.figure(figsize=(fig_w, fig_h))

    ax_x0 = label_w / fig_w
    ax_y0 = (ko_label_h + legend_h) / fig_h
    ax_w  = grid_w / fig_w
    ax_h  = grid_h / fig_h

    ax = fig.add_axes([ax_x0, ax_y0, ax_w, ax_h])
    ax.set_xlim(0, n_cols)
    ax.set_ylim(0, n_rows)
    ax.axis('off')

    # ── cells ───────────────────────────────────────────────────────────────
    for r in range(n_rows):
        row_y = n_rows - 1 - r
        if no_ref_row[r]:
            for c in range(n_cols):
                ax.add_patch(plt.Rectangle(
                    (c, row_y), 1, 1,
                    facecolor='#eeeeee', edgecolor='white', linewidth=0.5,
                    hatch='xxxx',
                ))
            continue
        for c in range(n_cols):
            val = matrix[r, c]
            fc = CATEGORY_COLORS[val]
            ax.add_patch(plt.Rectangle(
                (c, row_y), 1, 1,
                facecolor=fc, edgecolor='#dddddd', linewidth=0.5,
            ))

    # ── KO labels ───────────────────────────────────────────────────────────
    for c, (_, ko) in enumerate(ko_columns):
        label = ko_names.get(ko, ko) if ko_names else ko
        fig.text(
            ax_x0 + (c + 0.5) * cell_w / fig_w,
            ax_y0 - 0.012,
            label,
            ha='right', va='top', fontsize=22, rotation=45,
            transform=fig.transFigure,
        )

    # ── row labels ─────────────────────────────────────────────────────────
    for r, label in enumerate(row_labels):
        fig.text(
            ax_x0 + ax_w + 0.008,
            ax_y0 + (n_rows - 0.5 - r) * cell_h / fig_h,
            label,
            ha='left', va='center', fontsize=22,
            transform=fig.transFigure,
        )

    # ── pathway bars + labels ─────────────────────────────────────────────
    pw_span = {}
    for c, (pw, _) in enumerate(ko_columns):
        if pw not in pw_span:
            pw_span[pw] = [c, c]
        pw_span[pw][1] = c

    bar_y0_fig = ax_y0 + ax_h + 0.008
    bar_h_fig  = pw_bar_h / fig_h
    lbl_y0_fig = bar_y0_fig + bar_h_fig + 0.005

    for pw in pathway_order:
        start, end = pw_span[pw]
        color = pw_color[pw]
        bx = ax_x0 + start * cell_w / fig_w
        bw = (end - start + 1) * cell_w / fig_w

        fig.add_artist(plt.Rectangle(
            (bx, bar_y0_fig), bw, bar_h_fig,
            facecolor=color, transform=fig.transFigure,
            figure=fig, clip_on=False,
        ))
        mid_x = ax_x0 + (start + end + 1) / 2 * cell_w / fig_w
        pw_label = textwrap.fill(pw, width=12)
        fig.text(
            mid_x, lbl_y0_fig, pw_label,
            ha='center', va='bottom',
            fontsize=22, fontweight='bold', color=color,
            multialignment='center',
            transform=fig.transFigure,
        )
        if start > 0:
            sx = ax_x0 + start * cell_w / fig_w
            fig.add_artist(plt.Line2D(
                [sx, sx], [ax_y0, bar_y0_fig + bar_h_fig + 0.002],
                transform=fig.transFigure, figure=fig,
                color='#aaaaaa', linewidth=1.0, clip_on=False,
            ))

    # ── title ───────────────────────────────────────────────────────────────
    title_y = bar_y0_fig + bar_h_fig + pw_label_h / fig_h + 0.01
    fig.text(
        ax_x0 + ax_w / 2, title_y,
        title, ha='center', va='bottom',
        fontsize=26, fontweight='bold',
        transform=fig.transFigure,
    )

    # ── legend ─────────────────────────────────────────────────────────────
    legend_handles = [
        mpatches.Patch(facecolor=CATEGORY_COLORS[RARE_GAIN],
                        label=f'Rare gain (present, ≤{rare_threshold:.0%} of genus)'),
        mpatches.Patch(facecolor=CATEGORY_COLORS[RARE_LOSS],
                        label=f'Rare loss (absent, ≥{common_threshold:.0%} of genus)'),
        mpatches.Patch(facecolor=CATEGORY_COLORS[NORMAL_PRESENT], label='Normal presence'),
        mpatches.Patch(facecolor=CATEGORY_COLORS[NORMAL_ABSENT], edgecolor='#dddddd',
                        label='Normal absence'),
        mpatches.Patch(facecolor='#eeeeee', edgecolor='#999999', hatch='xxxx',
                        label='No genus reference (GlobDB)'),
    ]
    fig.legend(
        handles=legend_handles,
        loc='lower left',
        bbox_to_anchor=(ax_x0, 0.01),
        bbox_transform=fig.transFigure,
        frameon=False, fontsize=18, ncol=3,
    )

    plt.savefig(output_path, bbox_inches='tight', dpi=150)
    print(f"Saved: {output_path}")


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument('--summary',       required=True, help='summary.xlsx (reads its Genes sheet)')
    ap.add_argument('--mean-behavior', required=True, help='mean_behavior.xlsx (reads its first sheet)')
    ap.add_argument('--mean-sheet',    default='Sheet1', help='Sheet name in mean_behavior.xlsx (default: Sheet1)')
    ap.add_argument('--names', action='store_true', help='Show gene names instead of KO IDs where available')
    ap.add_argument('--genes-file', default='data/kegg_genes.csv')
    ap.add_argument('--ignore-pathway', action='append', default=[], metavar='PATHWAY')
    ap.add_argument('--include-all', action='store_true',
                     help='Include pathways absent in every genome (hidden by default)')
    ap.add_argument('--rare-threshold', type=float, default=0.10,
                     help='Genus-mean fraction at or below which presence counts as "rare gain" (default: 0.10)')
    ap.add_argument('--common-threshold', type=float, default=0.90,
                     help='Genus-mean fraction at or above which absence counts as "rare loss" (default: 0.90)')
    ap.add_argument('--output', default='rarity_grid.pdf')
    ap.add_argument('--title', default='Genome vs. genus background')
    args = ap.parse_args()

    ko_names = None
    if args.names:
        print(f"Loading KO names from {args.genes_file} ...")
        ko_names = load_ko_names(args.genes_file)

    print("Loading summary 'Genes' sheet ...")
    (genomes, species_labels, genus_labels, lineages,
     ko_columns_genes, values_genes) = load_numeric_grid(args.summary, 'Genes')
    print(f"  {len(genomes)} genomes")

    print("Loading mean_behavior sheet ...")
    _, _, _, _, ko_columns_mean, values_mean = load_numeric_grid(args.mean_behavior, args.mean_sheet)

    seen, pathway_order = set(), []
    for pw, _ in ko_columns_genes:
        if pw and pw not in seen:
            pathway_order.append(pw)
            seen.add(pw)

    auto_ignore = [] if args.include_all else find_empty_pathways(
        genomes, values_genes, ko_columns_genes, pathway_order)
    if auto_ignore:
        print(f"  Auto-ignoring (absent in all genomes): {', '.join(auto_ignore)}")
    ignore = DEFAULT_IGNORE | set(args.ignore_pathway) | set(auto_ignore)
    pathway_order = [p for p in pathway_order if p not in ignore]
    print(f"  Pathways shown ({len(pathway_order)}): {', '.join(pathway_order)}")

    print("Building rarity grid ...")
    ko_columns, matrix, row_labels, no_ref_row = build_category_grid(
        genomes, species_labels, genus_labels, lineages,
        ko_columns_genes, values_genes, ko_columns_mean, values_mean,
        pathway_order, args.rare_threshold, args.common_threshold,
    )

    # Sort genomes alphabetically by display label, keep no_ref_row aligned
    order = sorted(range(len(row_labels)), key=lambda i: row_labels[i].lower())
    row_labels = [row_labels[i] for i in order]
    matrix = matrix[order, :]
    no_ref_row = no_ref_row[order]

    print(f"  Grid: {matrix.shape[0]} genomes x {matrix.shape[1]} KOs")

    print("Plotting ...")
    plot_category_grid(ko_columns, matrix, row_labels, no_ref_row, args.title, args.output,
                        ko_names=ko_names, rare_threshold=args.rare_threshold,
                        common_threshold=args.common_threshold)


if __name__ == '__main__':
    main()
