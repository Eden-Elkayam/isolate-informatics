#!/usr/bin/env python3
"""
apply_representative_lineage.py

Fills in the species / genus / family / full_lineage columns of a pipeline
summary.xlsx using GTDB classifications from a representatives CSV (e.g. the
output of GTDB-Tk or a manual "representitives.csv"), matched by genome name.

Updates every sheet in the workbook that has the standard phylogeny header
columns (genome, species, genus, family, full_lineage), since export_excel.py
duplicates them onto the Pathways and Genes sheets.

Usage:
    python scripts/apply_representative_lineage.py \
        --summary runs/<run>/summary.xlsx \
        --representatives runs/<run>/representitives.csv \
        --output runs/<run>/summary.xlsx
"""

import argparse
import re
import openpyxl
import pandas as pd

PHYLO_COLS = ["species", "genus", "family", "full_lineage"]
RANK_PREFIXES = {"d": "domain", "p": "phylum", "c": "class",
                 "o": "order", "f": "family", "g": "genus", "s": "species"}


def load_lineages(representatives_path):
    """Return {genome: {species, genus, family, full_lineage}} keyed by
    genome name with the '.fa_assembly' suffix stripped."""
    rep = pd.read_csv(representatives_path)
    lineages = {}
    for _, row in rep.iterrows():
        genome = re.sub(r"\.fa_assembly$", "", row["User Genome"])
        classification = row["Classification"]
        if pd.isna(classification):
            continue
        ranks = {}
        for token in classification.split(";"):
            prefix, _, value = token.partition("__")
            if prefix in RANK_PREFIXES:
                ranks[RANK_PREFIXES[prefix]] = value
        lineages[genome] = {
            "species": ranks.get("species", ""),
            "genus": ranks.get("genus", ""),
            "family": ranks.get("family", ""),
            "full_lineage": classification,
        }
    return lineages


def update_sheet(ws, lineages):
    headers = [cell.value for cell in ws[1]]
    if "genome" not in headers or not all(c in headers for c in PHYLO_COLS):
        return 0

    genome_col = headers.index("genome") + 1
    phylo_cols = {col: headers.index(col) + 1 for col in PHYLO_COLS}

    updated = 0
    for row in ws.iter_rows(min_row=2):
        genome = row[genome_col - 1].value
        if genome in lineages:
            for col, lineage_value in lineages[genome].items():
                ws.cell(row=row[0].row, column=phylo_cols[col], value=lineage_value)
            updated += 1
    return updated


def main():
    parser = argparse.ArgumentParser(
        description="Apply representative-genome GTDB lineages to a summary.xlsx."
    )
    parser.add_argument("--summary", required=True, help="Input summary.xlsx")
    parser.add_argument("--representatives", required=True,
                        help="Representatives CSV with 'User Genome' and 'Classification' columns")
    parser.add_argument("--output", required=True, help="Output .xlsx path (can match --summary)")
    args = parser.parse_args()

    lineages = load_lineages(args.representatives)

    wb = openpyxl.load_workbook(args.summary)
    for sheet_name in wb.sheetnames:
        updated = update_sheet(wb[sheet_name], lineages)
        if updated:
            print(f"  {sheet_name}: updated {updated} row(s)")

    wb.save(args.output)
    print(f"Saved: {args.output}")


if __name__ == "__main__":
    main()
