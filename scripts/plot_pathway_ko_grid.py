#!/usr/bin/env python3
"""
plot_pathway_ko_grid.py

Creates a KO × genome presence grid for pathways that have sanity-check expectations.

Columns  = individual KOs from pathway definitions, grouped under colored pathway bars.
           Only pathways with at least one species expectation are shown.
Rows     = genomes from the summary Excel.
Colors:
  solid (pathway color)  — KO is present in the genome
  light pink / dashed    — KO is absent and expected (genome expected to have
                           this pathway; the missing KO contributes to failure)
  light grey             — KO is absent, not expected for this genome

OR-group logic for "expected-but-missing":
  If a pathway fails and the missing KO is part of an OR-group (alternatives),
  ALL alternatives in that OR-group are coloured pink (any one would satisfy it).
  If one alternative IS present, the OR-group passes and nothing is pink.

Usage
-----
python plot_pathway_ko_grid.py \\
    --expectations  c1_informatics/data/species_pathway_expectations.csv \\
    --summary       runs/dram_examples/summary.xlsx \\
    --pathway-defs  c1_informatics/data/pathway_definitions.csv \\
    --output        runs/dram_examples/ko_grid.pdf \\
    --title         "DRAM examples — pathway KO grid"
"""

import argparse
import csv
import re
import textwrap
from collections import defaultdict
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl


# ─────────────────────────── expression parser ───────────────────────────────

def _tokenize(expr):
    return re.findall(r'K\d+|\(|\)|AND|OR', expr)


def _parse_or(tok, pos):
    left, pos = _parse_and(tok, pos)
    children = [left]
    while pos < len(tok) and tok[pos] == 'OR':
        pos += 1
        right, pos = _parse_and(tok, pos)
        children.append(right)
    return (('OR', children) if len(children) > 1 else children[0]), pos


def _parse_and(tok, pos):
    left, pos = _parse_atom(tok, pos)
    children = [left]
    while pos < len(tok) and tok[pos] == 'AND':
        pos += 1
        right, pos = _parse_atom(tok, pos)
        children.append(right)
    return (('AND', children) if len(children) > 1 else children[0]), pos


def _parse_atom(tok, pos):
    t = tok[pos]
    if t == '(':
        node, pos = _parse_or(tok, pos + 1)
        assert tok[pos] == ')', f"Expected ')' at pos {pos}"
        return node, pos + 1
    if re.fullmatch(r'K\d+', t):
        return ('KO', t), pos + 1
    raise ValueError(f"Unexpected token '{t}' at pos {pos}")


def parse_expression(expr):
    tok = _tokenize(expr)
    if not tok:
        return None
    node, _ = _parse_or(tok, 0)
    return node


def eval_node(node, present):
    """Evaluate expression; present is a set of KO strings."""
    kind = node[0]
    if kind == 'KO':
        return node[1] in present
    if kind == 'AND':
        return all(eval_node(c, present) for c in node[1])
    if kind == 'OR':
        return any(eval_node(c, present) for c in node[1])


def all_kos(node):
    if node[0] == 'KO':
        return frozenset([node[1]])
    return frozenset().union(*(all_kos(c) for c in node[1]))


def failed_kos(node, present):
    """
    KOs whose absence contributed to pathway failure.
    AND → collect failures from each child.
    OR  → if OR fails entirely, return all KOs in that OR-group.
          if OR passes, return empty (the alternative was satisfied).
    """
    kind = node[0]
    if kind == 'KO':
        return frozenset([node[1]]) if node[1] not in present else frozenset()
    if kind == 'AND':
        result = frozenset()
        for c in node[1]:
            result |= failed_kos(c, present)
        return result
    if kind == 'OR':
        if eval_node(node, present):
            return frozenset()
        return all_kos(node)  # OR failed: all KOs in the group are red


def ordered_kos(node):
    """KOs in left-to-right order as they appear in the expression."""
    if node[0] == 'KO':
        return [node[1]]
    seen, result = set(), []
    for c in node[1]:
        for ko in ordered_kos(c):
            if ko not in seen:
                result.append(ko)
                seen.add(ko)
    return result


# ─────────────────────────── species matching ────────────────────────────────

def _norm(name):
    """(genus_lower, first_species_word_lower) ignoring clade markers."""
    name = re.sub(r'^s__', '', str(name)).strip()
    parts = re.split(r'[\s_]+', name)
    parts = [p for p in parts if p and not re.fullmatch(r'[A-Z]', p)]
    g = parts[0].lower() if parts else ''
    s = parts[1].lower() if len(parts) > 1 else ''
    return g, s


def build_species_map(expectations_path):
    """Returns dict: (genus, species_word) -> set of pathway names."""
    mapping = defaultdict(set)
    with open(expectations_path) as f:
        for row in csv.DictReader(f):
            mapping[_norm(row['species'])].add(row['pathway'])
    return dict(mapping)


def expected_pathways_for(species_label, species_map):
    """Return set of expected pathways for a GTDB/custom species label."""
    key = _norm(species_label)
    if key in species_map:
        return species_map[key]
    # Fallback: genus-only match (only if unique)
    genus = key[0]
    matches = [p for (g, _), p in species_map.items() if g == genus]
    if len(matches) == 1:
        return matches[0]
    return set()


# ─────────────────────────── load inputs ─────────────────────────────────────

def load_ko_names(path):
    """dict: KO -> name from kegg_genes.csv 'name' column. KOs without a name are omitted (caller falls back to KO ID)."""
    names = {}
    with open(path) as f:
        for row in csv.DictReader(f):
            ko = row.get('KO', '').strip()
            name = row.get('name', '').strip()
            if ko and name:
                names[ko] = name
    return names


def load_pathway_defs(path):
    """dict: pathway -> {'tree': node, 'kos': [ordered KO list]}"""
    defs = {}
    with open(path) as f:
        for row in csv.DictReader(f):
            name = row['pathway'].strip()
            tree = parse_expression(row['expression'].strip())
            if tree is None:
                continue
            defs[name] = {'tree': tree, 'kos': ordered_kos(tree)}
    return defs


def load_summary(summary_path):
    """
    Reads the 'Genes' sheet from summary.xlsx.

    Sheet layout (one row per genome):
      col 0 : genome_id
      col 1 : species label
      col 2 : genus
      col 3 : family
      col 4 : full_lineage
      col 5+: pathway KO hits (comma-separated KO IDs or 'None')

    Returns
    -------
    genomes        : list of genome IDs in row order
    species_labels : {genome_id: species_label}
    ko_hits        : {(genome_id, pathway): set of KO strings present}
    pathway_cols   : ordered list of pathway names (from header)
    """
    wb = openpyxl.load_workbook(summary_path, read_only=True, data_only=True)
    if 'Genes' not in wb.sheetnames:
        raise ValueError(f"No 'Genes' sheet found in {summary_path}.\n"
                         "The summary must be generated by print_summary.sh.")
    ws = wb['Genes']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Genes sheet is empty.")

    header = rows[0]
    # Columns 0-4 are genome/taxonomy fields; 5+ are pathway columns
    META_COLS = 5
    pathway_cols = [str(h).strip() for h in header[META_COLS:] if h is not None]

    genomes = []
    species_labels = {}
    ko_hits = {}

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        genome_id = str(row[0]).strip()
        species = str(row[1]).strip() if row[1] else genome_id
        genomes.append(genome_id)
        species_labels[genome_id] = species

        for i, pathway in enumerate(pathway_cols):
            cell = row[META_COLS + i]
            val = str(cell).strip() if cell else ''
            kos = set()
            if val and val.lower() != 'none':
                kos = {k.strip() for k in val.split(',') if k.strip().startswith('K')}
            ko_hits[(genome_id, pathway)] = kos

    return genomes, species_labels, ko_hits, pathway_cols


# Pathways always excluded from the plot regardless of data
DEFAULT_IGNORE = set()

# ─────────────────────────── build grid ──────────────────────────────────────

ABSENT      = 0
PRESENT     = 1
MISSING_EXP = 2


def build_grid(genomes, species_labels, ko_hits, pathway_defs,
               species_map, expected_pathway_order):
    """
    Returns
    -------
    ko_columns : list of (pathway, ko) in display order
    matrix     : ndarray (n_genomes × n_kos), values in {ABSENT, PRESENT, MISSING_EXP}
    row_labels : display label per genome row
    """
    # Build ordered KO columns — all KOs for each pathway, in definition order.
    # A KO shared by multiple pathways appears once per pathway.
    ko_columns = []
    for pathway in expected_pathway_order:
        if pathway not in pathway_defs:
            continue
        for ko in pathway_defs[pathway]['kos']:
            ko_columns.append((pathway, ko))

    n_rows = len(genomes)
    n_cols = len(ko_columns)
    matrix = np.zeros((n_rows, n_cols), dtype=int)
    row_labels = []

    for r, gid in enumerate(genomes):
        sp = species_labels.get(gid, gid)
        # Clean up display label
        label = re.sub(r'^s__', '', sp).strip()
        row_labels.append(label)

        expected = expected_pathways_for(sp, species_map)

        # Determine which KOs are "red" across all expected pathways
        red_kos = set()
        for pathway in expected:
            if pathway not in pathway_defs:
                continue
            tree = pathway_defs[pathway]['tree']
            present = ko_hits.get((gid, pathway), set())
            if not eval_node(tree, present):
                red_kos |= failed_kos(tree, present)

        for c, (col_pathway, ko) in enumerate(ko_columns):
            present = ko_hits.get((gid, col_pathway), set())
            if ko in present:
                matrix[r, c] = PRESENT
            elif ko in red_kos:
                matrix[r, c] = MISSING_EXP
            else:
                matrix[r, c] = ABSENT

    return ko_columns, matrix, row_labels


# ─────────────────────────── plot ────────────────────────────────────────────

_PALETTE = [
    '#2ca02c', '#ff7f0e', '#1f77b4', '#9467bd', '#8c564b',
    '#e377c2', '#d62728', '#17becf', '#bcbd22', '#7f7f7f',
    '#aec7e8', '#ffbb78', '#98df8a', '#f7b6d2', '#c5b0d5',
    '#c49c94', '#dbdb8d', '#9edae5', '#ff9896', '#c7c7c7',
]


def plot_grid(ko_columns, matrix, row_labels, title, output_path, ko_names=None):
    n_rows, n_cols = matrix.shape

    # Pathway order + colors
    pathway_order, seen = [], set()
    for pw, _ in ko_columns:
        if pw not in seen:
            pathway_order.append(pw)
            seen.add(pw)
    pw_color = {p: _PALETTE[i % len(_PALETTE)] for i, p in enumerate(pathway_order)}

    # Geometry
    cell_w  = 1.1   # wider than tall so rotated labels have more horizontal room
    cell_h  = 0.55
    label_w = max((len(s) for s in row_labels), default=10) * 0.18 + 0.5
    ko_label_h = 5.0   # rotated KO labels below cells
    pw_bar_h   = 0.45  # pathway bar height (inches)
    pw_label_h = 3.0   # pathway label area above bar (2-line headers need more)
    legend_h   = 0.7

    grid_w = n_cols * cell_w
    grid_h = n_rows * cell_h

    fig_w = max(14, label_w + grid_w + 0.5)
    fig_h = ko_label_h + grid_h + pw_bar_h + pw_label_h + legend_h + 0.3

    fig = plt.figure(figsize=(fig_w, fig_h))

    # Axes origin in figure-fraction units
    ax_x0 = label_w / fig_w
    ax_y0 = (ko_label_h + legend_h) / fig_h
    ax_w  = grid_w / fig_w
    ax_h  = grid_h / fig_h

    ax = fig.add_axes([ax_x0, ax_y0, ax_w, ax_h])
    ax.set_xlim(0, n_cols)
    ax.set_ylim(0, n_rows)
    ax.axis('off')

    # ── cells ──────────────────────────────────────────────────────────────
    for r in range(n_rows):
        for c in range(n_cols):
            val = matrix[r, c]
            pw  = ko_columns[c][0]
            row_y = n_rows - 1 - r  # top row = r=0

            if val == PRESENT:
                fc, ec, lw, ls = pw_color[pw], 'white', 0.5, '-'
            elif val == MISSING_EXP:
                fc, ec, lw, ls = '#fce4e4', '#d9534f', 1.2, '--'
            else:
                fc, ec, lw, ls = '#f0f0f0', 'white', 0.5, '-'

            ax.add_patch(plt.Rectangle(
                (c, row_y), 1, 1,
                facecolor=fc, edgecolor=ec, linewidth=lw, linestyle=ls
            ))

    # ── KO labels (below grid) ─────────────────────────────────────────────
    for c, (_, ko) in enumerate(ko_columns):
        label = ko_names.get(ko, ko) if ko_names else ko
        fig.text(
            ax_x0 + (c + 0.5) * cell_w / fig_w,
            ax_y0 - 0.012,
            label,
            ha='right', va='top', fontsize=22, rotation=45,
            transform=fig.transFigure
        )

    # ── row labels (right of grid) ─────────────────────────────────────────
    for r, label in enumerate(row_labels):
        fig.text(
            ax_x0 + ax_w + 0.008,
            ax_y0 + (n_rows - 0.5 - r) * cell_h / fig_h,
            label,
            ha='left', va='center', fontsize=22,
            transform=fig.transFigure
        )

    # ── pathway bars + labels (above grid) ────────────────────────────────
    # Compute span of each pathway
    pw_span = {}
    for c, (pw, _) in enumerate(ko_columns):
        if pw not in pw_span:
            pw_span[pw] = [c, c]
        pw_span[pw][1] = c

    bar_y0_fig = ax_y0 + ax_h + 0.008          # bottom of bar strip
    bar_h_fig  = pw_bar_h / fig_h
    lbl_y0_fig = bar_y0_fig + bar_h_fig + 0.005

    for pw in pathway_order:
        start, end = pw_span[pw]
        color = pw_color[pw]
        bx = ax_x0 + start * cell_w / fig_w
        bw = (end - start + 1) * cell_w / fig_w

        fig.add_artist(plt.Rectangle(
            (bx, bar_y0_fig), bw, bar_h_fig,
            facecolor=color, transform=fig.transFigure,
            figure=fig, clip_on=False
        ))

        mid_x = ax_x0 + (start + end + 1) / 2 * cell_w / fig_w
        pw_label = textwrap.fill(pw, width=12)
        fig.text(
            mid_x, lbl_y0_fig, pw_label,
            ha='center', va='bottom',
            fontsize=22, fontweight='bold', color=color,
            multialignment='center',
            transform=fig.transFigure
        )

        # vertical separator before this pathway
        if start > 0:
            sx = ax_x0 + start * cell_w / fig_w
            fig.add_artist(plt.Line2D(
                [sx, sx],
                [ax_y0, bar_y0_fig + bar_h_fig + 0.002],
                transform=fig.transFigure, figure=fig,
                color='#aaaaaa', linewidth=1.0, clip_on=False
            ))

    # ── title ──────────────────────────────────────────────────────────────
    title_y = bar_y0_fig + bar_h_fig + pw_label_h / fig_h + 0.01
    fig.text(
        ax_x0 + ax_w / 2, title_y,
        title, ha='center', va='bottom',
        fontsize=26, fontweight='bold',
        transform=fig.transFigure
    )

    # ── legend ─────────────────────────────────────────────────────────────
    legend_handles = [
        mpatches.Patch(facecolor='#2ca02c', label='Present'),
        mpatches.Patch(facecolor='#fce4e4', edgecolor='#d9534f',
                       linestyle='--', linewidth=1.5, label='Missing (expected)'),
        mpatches.Patch(facecolor='#f0f0f0', label='Not expected'),
    ]
    fig.legend(
        handles=legend_handles,
        loc='lower left',
        bbox_to_anchor=(ax_x0, 0.01),
        bbox_transform=fig.transFigure,
        frameon=False, fontsize=20, ncol=3
    )

    plt.savefig(output_path, bbox_inches='tight', dpi=150)
    print(f"Saved: {output_path}")


# ─────────────────────────── main ────────────────────────────────────────────

def find_empty_pathways(genomes, ko_hits, pathway_order):
    """Return pathways that have no KO hits in any genome."""
    return [p for p in pathway_order
            if not any(ko_hits.get((gid, p), set()) for gid in genomes)]


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument('--expectations', required=True,
                    help='species_pathway_expectations.csv')
    ap.add_argument('--summary',      required=True,
                    help='summary.xlsx from print_summary.sh')
    ap.add_argument('--pathway-defs', required=True,
                    help='pathway_definitions.csv')
    ap.add_argument('--names', action='store_true',
                    help='Show gene names instead of KO IDs where available')
    ap.add_argument('--genes-file', default='data/kegg_genes.csv',
                    help='Path to kegg_genes.csv (default: data/kegg_genes.csv)')
    ap.add_argument('--ignore-pathway', action='append', default=[], metavar='PATHWAY',
                    help='Pathway name to exclude (repeat for multiple)')
    ap.add_argument('--include-all', action='store_true',
                    help='Include pathways absent in every genome (ignore list still applies)')
    ap.add_argument('--output',  default='ko_grid.pdf',
                    help='Output file path (PDF or PNG)')
    ap.add_argument('--title',   default='Pathway KO Presence Grid',
                    help='Plot title')
    args = ap.parse_args()

    ko_names = None
    if args.names:
        print(f"Loading KO names from {args.genes_file} ...")
        ko_names = load_ko_names(args.genes_file)

    print("Loading pathway definitions ...")
    pathway_defs = load_pathway_defs(args.pathway_defs)

    print("Loading expectations ...")
    species_map = build_species_map(args.expectations)
    all_expected = set()
    for pw_set in species_map.values():
        all_expected |= pw_set

    # Keep definition order, filter to expected pathways only
    expected_pathway_order = [p for p in pathway_defs if p in all_expected]

    print("Loading summary Excel ...")
    genomes, species_labels, ko_hits, _ = load_summary(args.summary)
    print(f"  {len(genomes)} genomes")

    # Auto-ignore pathways absent in all genomes, plus defaults and user-specified
    auto_ignore = [] if args.include_all else find_empty_pathways(genomes, ko_hits, expected_pathway_order)
    ignore = DEFAULT_IGNORE | set(args.ignore_pathway) | set(auto_ignore)
    if auto_ignore:
        print(f"  Auto-ignoring (absent in all genomes): {', '.join(auto_ignore)}")
    if DEFAULT_IGNORE:
        print(f"  Always ignoring: {', '.join(DEFAULT_IGNORE)}")
    if args.ignore_pathway:
        print(f"  User-ignoring: {', '.join(args.ignore_pathway)}")
    expected_pathway_order = [p for p in expected_pathway_order if p not in ignore]
    print(f"  Pathways shown ({len(expected_pathway_order)}): "
          f"{', '.join(expected_pathway_order)}")

    # Sort genomes alphabetically by species label
    genomes = sorted(genomes, key=lambda g: re.sub(r'^s__', '', species_labels.get(g, g)).lower())

    print("Building grid ...")
    ko_columns, matrix, row_labels = build_grid(
        genomes, species_labels, ko_hits,
        pathway_defs, species_map, expected_pathway_order
    )
    print(f"  Grid: {len(genomes)} genomes × {len(ko_columns)} KOs")

    print("Plotting ...")
    plot_grid(ko_columns, matrix, row_labels, args.title, args.output, ko_names=ko_names)


if __name__ == '__main__':
    main()
